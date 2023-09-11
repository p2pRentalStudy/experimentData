from basicImports import *
import requests
import random
import string
from os import walk
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
import os
from natsort import natsorted
from openpyxl.styles import Alignment

# from PIL import Image, ImageOps




studiedApps = {}
# studiedApps['Turo'] = 1
studiedApps['GerAroundEurope'] = 1
# studiedApps['GetAround'] = 1

# jvs4v1q5

#ownerimages
targetFolder = '/home/hakhan/Google Drive/p2pCarRentalProject/AnalysisScripts/IntermediateData/ownerImages/APPNAME/'
for platform in studiedApps:
    tempFolder = targetFolder.replace('APPNAME',platform)
    csvPath = '/home/hakhan/Google Drive/p2pCarRentalProject/AnalysisScripts/IntermediateData/ownerImages/'+platform+'.xlsx'

    wb = Workbook()
    ws = wb.active


    # writer.writerow(['ID', 'Img' 'Gender', 'Age', 'Race'])
    

    #read all img files and write rows
    filenames = next(walk(tempFolder), (None, None, []))[2]  # [] if no file
    filenames.sort()

    print(len(filenames))
    # time.sleep(10000)
    index = 1

    height = 120
    width = 120

    ws.cell(row=1, column=1, value='ID')
    ws.cell(row=1, column=2, value='Image')
    # ws.cell(row=1, column=3, value='Name')
    ws.cell(row=1, column=3, value='P#')#people count
    ws.cell(row=1, column=4, value='Gender') #m=male, f= female, u=unknown pic, g=image of rental company
    ws.cell(row=1, column=5, value='Age')
    ws.cell(row=1, column=6, value='Race')

    for fn in filenames:
        imgFile = tempFolder+fn
        img = Image(imgFile)
        objectID = fn.split('.jpg')[0]
        img.width = width
        img.height = height

        ws.row_dimensions[index+1].height = 90
        # ws.row_dimensions[index+1].width = width
        for col in range(1,10):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
            # ws.column_dimensions[col_letter].height = height
        # ws.add_text(objectID, anchor='A'+str(index+1))
        ws.add_image(img, anchor='B'+str(index+1))
        ws.cell(row=index+1, column=1, value=objectID)
        ws['A'+str(index+1)].alignment = Alignment(wrap_text=True)

        # print('A'+str(index+1))
        index += 1
        # if index > 10:
        #     break
    

    wb.save(csvPath)




    
    